from collections import Counter
from matplotlib import pyplot as grap
import openpyxl
import re

wookbook = openpyxl.load_workbook("FINAL_DATA.xlsx")
worksheet = wookbook.active

def DIAGRAM_1():
    WORDS = ""
    LIST = []
    LIST.clear()
    for i in range(0, worksheet.max_row):
        for j in worksheet.iter_cols(1, 1):
            s = j[i].value
            LIST.append(s)
    WORDS = ' '.join(LIST)
    shortword = re.compile(r'\W*\b\w{1,3}\b')
    WORDS = shortword.sub('', WORDS).lower().replace('–', '')
    X = Counter(WORDS.split())
    X = X.most_common(10)
    labels, values = zip(*X)
    grap.pie(values, labels=labels)
    grap.title('Слова встречающиеся чаще всего')
    grap.show()

def DIAGRAM_2_additional_statistics():
    new_list = []
    for i in range(0, worksheet.max_row):
        for j in worksheet.iter_cols(2):
            s = j[i].value
            s = re.sub('\s+', '', s)
            if s.isnumeric():
                new_list.append(s)
    data = list(map(int, filter(lambda x: int(x) > 1000, new_list)))
    X = Counter(data)
    X = X.most_common(200)
    X = sorted(X, key=lambda x: x[0], reverse=True)
    labels, values = zip(*X)
    colors = ['red' if (bar == max(data)) else 'pink' for bar in values]
    grap.bar(labels, values, color=colors)
    grap.title('Количество книг по справочной литературе по годам')
    grap.xlabel('Год')
    grap.ylabel('Количество')
    grap.grid(color = 'blue', alpha = 0.3, linestyle = '-', linewidth = 0.5)
    grap.show()

DIAGRAM_1()
DIAGRAM_2_additional_statistics()


