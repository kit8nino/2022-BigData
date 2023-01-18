import openpyxl
import re
from collections import Counter
from matplotlib import pyplot as grap

wookbook = openpyxl.load_workbook("books.xlsx")
worksheet = wookbook.active

def diagram_years():
    new_list = []
    for i in range(0, worksheet.max_row):
        for j in worksheet.iter_cols(2):
            s = j[i].value
            s = re.sub('\s+', '', s)
            if s.isnumeric():
                new_list.append(s)
    words = filter(lambda x: int(x) > 1000, new_list)
    data = list(map(int, words))
    k = Counter(data)
    k = k.most_common(200)
    k = sorted(k, key=lambda x: x[0], reverse=True)
    labels, values = zip(*k)
    colors = ['red' if (bar == max(data)) else 'blue' for bar in values]
    grap.bar(labels, values, color=colors)
    grap.title('Количество книг по религии и культуре по годам')
    grap.xlabel('Год')
    grap.ylabel('Количество')
    grap.grid(color='red', alpha=0.3, linestyle='-', linewidth=0.5)
    grap.show()

def diagram_words():
    words = ""
    list = []
    list.clear()
    for i in range(0, worksheet.max_row):
        for j in worksheet.iter_cols(1, 1):
            s = j[i].value
            list.append(s)
    words = ' '.join(list)
    shortword = re.compile(r'\W\b\w{1,3}\b')
    words = shortword.sub('', words)
    words = words.lower()
    words = words.replace('–', '')
    k = Counter(words.split())
    k = k.most_common(7)
    labels, values = zip(*k)
    grap.pie(values, labels=labels)
    grap.title('Слова встречающиеся чаще всего')
    grap.show()

diagram_years()
diagram_words()